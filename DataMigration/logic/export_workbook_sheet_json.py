from __future__ import annotations

import csv
import json
import re
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


WORKBOOK_PATH = Path(r"c:\Projects\Meter App\legacy-ud\workbook-exports\macro-free\Utility Dash 9 Mar decrypted working copy (2) 2026.xlsx")
BASE_DIR = Path(r"c:\Projects\Meter App\DataMigration")
INPUT_DIR = BASE_DIR / "inputs" / "sheet-json"
NORMALIZED_DIR = BASE_DIR / "outputs" / "sheet-normalized"
REVIEW_DIR = BASE_DIR / "outputs" / "reviews"

SUPPORT_SHEETS = {
    'CheckList', 'ReadableReport', 'BuildingReport', 'BuildingReportTemplate', 'Variable_List',
    'ReadableReportTemplate (3)', 'ReadableReportTemplate (2)', 'Home', 'Help', 'Settings', 'Tariffs',
    'UnitHistory', 'ElecBreakDown', 'WaterBreakdown', 'ReadableReportTemplate', 'ReadableReportTemplate Water',
    'Invoice', 'WCU Output', 'BCM Output', 'Sheet2'
}

SCHEME_ALIASES = {
    'the azores': 'Azores',
    'genesis': 'Genisis on Fairmount',
    'bonifay court': 'Bonifay',
    'vilino glen': 'VILLINO GLEN',
    'l montagne': 'La Montagne',
    'lmontagne': 'La Montagne',
    "l'montagne": 'La Montagne',
    'rivonia gate': 'Rivonia Gates',
    'queensgate': 'QueensGate',
}


@dataclass
class MonthColumn:
    column_index: int
    display_label: str | None
    reading_date_raw: Any
    tariff_table: str | None


def slugify(value: str) -> str:
    slug = re.sub(r'[^a-z0-9]+', '-', value.strip().lower())
    return slug.strip('-') or 'unknown'


def normalize_token(value: Any) -> str:
    return re.sub(r'[^a-z0-9]+', ' ', str(value or '').strip().lower()).strip()


def canonical_scheme_name(sheet_name: str) -> str:
    normalized = normalize_token(sheet_name)
    return SCHEME_ALIASES.get(normalized, sheet_name.strip())


def serialize_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat(sep=' ')
    return value


def parse_text_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text or text == '#########':
        return None
    normalized = text.replace(' ', '').replace(',', '')
    try:
        return float(normalized)
    except ValueError:
        return None


def parse_reading_date(raw_value: Any, display_label: str | None) -> tuple[str | None, str | None]:
    if isinstance(raw_value, datetime):
        return raw_value.date().isoformat(), None

    if raw_value is None:
        return None, 'missing-date'

    text = str(raw_value).strip()
    if not text or text == '#########':
        return None, 'placeholder-date'

    formats = ['%Y-%m', '%Y/%m/%d', '%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y', '%Y-%m-%d %H:%M:%S']
    for fmt in formats:
        try:
            parsed = datetime.strptime(text, fmt)
            if fmt == '%Y-%m':
                return parsed.strftime('%Y-%m-01'), None
            return parsed.date().isoformat(), None
        except ValueError:
            continue

    if display_label:
        for fmt in ['%b-%y', '%B-%y']:
            try:
                parsed = datetime.strptime(display_label, fmt)
                return parsed.strftime('%Y-%m-01'), 'derived-from-label'
            except ValueError:
                continue

    return None, 'unparseable-date'


def get_settings(ws) -> dict[str, Any]:
    settings = {
        'vat_percent': None,
        'ignore_negative_common_property_amounts': None,
        'tolerance_months': None,
        'tolerance_percentage': None,
    }
    for row in ws.iter_rows(min_row=1, max_row=40, min_col=1, max_col=4, values_only=True):
        row_values = list(row)
        if len(row_values) < 3:
            continue
        description = row_values[2]
        value = row_values[1]
        if description == 'Value' and row_values[1] == 'Description':
            continue
        if description == 'Ignore Negative Common Proper Amounts':
            settings['ignore_negative_common_property_amounts'] = bool(value)
        elif description == 'Error Tolerance Percentage':
            settings['tolerance_percentage'] = parse_text_number(value)
        elif description == 'Number of Months to compare average':
            settings['tolerance_months'] = int(parse_text_number(value) or 0)
        elif row_values[1] == 'VAT':
            settings['vat_percent'] = parse_text_number(description)
    return settings


def get_month_columns(ws) -> list[MonthColumn]:
    columns: list[MonthColumn] = []
    column_index = 5
    while True:
        raw_date = ws.cell(2, column_index).value
        display_label = ws.cell(1, column_index).value
        tariff_table = ws.cell(3, column_index).value
        if raw_date in (None, '') and display_label in (None, ''):
            break
        columns.append(MonthColumn(
            column_index=column_index,
            display_label=str(display_label).strip() if display_label is not None else None,
            reading_date_raw=raw_date,
            tariff_table=str(tariff_table).strip() if tariff_table is not None else None,
        ))
        column_index += 1
    return columns


def find_row_by_label(ws, label: str, max_rows: int) -> int | None:
    target = normalize_token(label)
    for row_index in range(1, max_rows + 1):
        cell_value = normalize_token(ws.cell(row_index, 1).value)
        if cell_value == target:
            return row_index
    return None


def classify_meter_type(label: str) -> str:
    normalized = str(label or '').replace('*', '').strip().upper()
    if not normalized:
        return 'UNIT'
    if normalized.startswith('BUL') or 'BULK' in normalized:
        return 'BULK'
    if normalized.startswith('COM'):
        return 'COMMON'
    common_tokens = ['GATE', 'GUARD', 'STORE', 'PUMP', 'POOL', 'LIFT', 'BOREHOLE', 'CLUBHOUSE', 'OFFICE', 'LAUNDRY', 'LIGHT', 'LOBBY', 'SEWER', 'GARDEN', 'SPRINKLER', 'PUBLIC']
    if any(token in normalized for token in common_tokens):
        return 'COMMON'
    return 'UNIT'


def extract_non_empty_rows(ws) -> list[dict[str, Any]]:
    rows = []
    for row_index in range(1, ws.max_row + 1):
        cells = []
        for column_index in range(1, ws.max_column + 1):
            value = ws.cell(row_index, column_index).value
            if value in (None, ''):
                continue
            cells.append({
                'column_index': column_index,
                'column_letter': get_column_letter(column_index),
                'value': serialize_value(value),
            })
        if cells:
            rows.append({
                'row_index': row_index,
                'cells': cells,
            })
    return rows


def extract_building_sheet(ws, settings: dict[str, Any]) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    review_flags: list[dict[str, Any]] = []
    max_rows = ws.max_row
    header_scan_limit = min(max_rows, 250)
    electricity_header = find_row_by_label(ws, 'ELECTRICITY', header_scan_limit)
    water_header = find_row_by_label(ws, 'WATER', header_scan_limit)
    month_columns = get_month_columns(ws)

    normalized_months = []
    for month in month_columns:
        reading_date, reading_date_note = parse_reading_date(month.reading_date_raw, month.display_label)
        month_record = {
            'column_index': month.column_index,
            'display_label': month.display_label,
            'reading_date_raw': serialize_value(month.reading_date_raw),
            'reading_date': reading_date,
            'reading_date_note': reading_date_note,
            'tariff_table': month.tariff_table,
        }
        normalized_months.append(month_record)

        if not month.display_label:
            review_flags.append({
                'severity': 'medium',
                'type': 'missing-month-display-label',
                'column_index': month.column_index,
                'message': 'Month column is missing a display label.'
            })
        if reading_date_note in {'missing-date', 'placeholder-date', 'unparseable-date'}:
            review_flags.append({
                'severity': 'high',
                'type': 'unresolved-reading-date',
                'column_index': month.column_index,
                'display_label': month.display_label,
                'reading_date_raw': serialize_value(month.reading_date_raw),
                'message': f'Reading date could not be resolved for month column {month.column_index}.'
            })

    charge_modes = []
    for row_index in range(4, 12):
        label = ws.cell(row_index, 1).value
        distribution_mode = ws.cell(row_index, 3).value
        if label:
            charge_modes.append({
                'source_row': row_index,
                'charge_name': str(label).strip(),
                'distribution_mode': str(distribution_mode).strip() if distribution_mode is not None else None,
            })
        else:
            review_flags.append({
                'severity': 'medium',
                'type': 'missing-charge-heading',
                'row_index': row_index,
                'message': 'Expected charge heading row is blank.'
            })

    if electricity_header is None:
        review_flags.append({
            'severity': 'high',
            'type': 'missing-electricity-header',
            'message': 'The sheet does not contain an ELECTRICITY section header.'
        })
        return {
            'month_columns': normalized_months,
            'charge_modes': charge_modes,
            'sections': {
                'electricity_header_row': None,
                'water_header_row': water_header,
            },
            'electricity_rows': [],
        }, review_flags

    electricity_rows = []
    seen_labels = Counter()
    current_row = electricity_header + 1
    blank_streak = 0
    while current_row <= max_rows:
        label = ws.cell(current_row, 1).value
        if label is None:
            blank_streak += 1
            if blank_streak >= 12:
                break
            current_row += 1
            continue

        blank_streak = 0
        label_text = str(label).strip()
        upper_label = label_text.upper()

        if upper_label == 'WATER':
            break
        if upper_label in {'ELECTRICITY', 'UNIT'}:
            current_row += 1
            continue

        seen_labels[label_text] += 1
        meter_type = classify_meter_type(label_text)
        prepaid_marker = ws.cell(current_row, 2).value
        pq_factor = parse_text_number(ws.cell(current_row, 3).value)
        readings = []
        placeholder_count = 0

        for month in normalized_months:
            raw_value = ws.cell(current_row, month['column_index']).value
            if raw_value == '#########':
                placeholder_count += 1
            reading_value = parse_text_number(raw_value)
            if raw_value in (None, '') and reading_value is None:
                continue
            readings.append({
                'column_index': month['column_index'],
                'reading_label': month['display_label'],
                'reading_date_raw': month['reading_date_raw'],
                'reading_date': month['reading_date'],
                'reading_date_note': month['reading_date_note'],
                'tariff_table': month['tariff_table'],
                'reading_value_raw': serialize_value(raw_value),
                'reading_value': reading_value,
            })

        if not readings:
            review_flags.append({
                'severity': 'medium',
                'type': 'row-with-no-readings',
                'row_index': current_row,
                'legacy_label': label_text,
                'message': 'Electricity row has no captured reading values.'
            })

        if placeholder_count > 0:
            review_flags.append({
                'severity': 'high',
                'type': 'placeholder-reading-values',
                'row_index': current_row,
                'legacy_label': label_text,
                'count': placeholder_count,
                'message': 'Electricity row contains placeholder reading values.'
            })

        electricity_rows.append({
            'source_row': current_row,
            'legacy_label': label_text,
            'meter_type': meter_type,
            'prepaid_marker': serialize_value(prepaid_marker),
            'pq_factor': pq_factor,
            'readings': readings,
        })
        current_row += 1

    for legacy_label, count in seen_labels.items():
        if count > 1:
            review_flags.append({
                'severity': 'high',
                'type': 'duplicate-legacy-label',
                'legacy_label': legacy_label,
                'count': count,
                'message': 'Legacy label appears more than once in the ELECTRICITY section.'
            })

    normalized_sheet = {
        'month_columns': normalized_months,
        'charge_modes': charge_modes,
        'sections': {
            'electricity_header_row': electricity_header,
            'water_header_row': water_header,
        },
        'electricity_rows': electricity_rows,
    }
    return normalized_sheet, review_flags


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=True), encoding='utf-8')


def write_review_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        'sheet_name', 'severity', 'type', 'message', 'row_index', 'column_index',
        'legacy_label', 'display_label', 'reading_date_raw', 'count'
    ]
    with path.open('w', newline='', encoding='utf-8') as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    INPUT_DIR.mkdir(parents=True, exist_ok=True)
    NORMALIZED_DIR.mkdir(parents=True, exist_ok=True)
    REVIEW_DIR.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(WORKBOOK_PATH, data_only=False, read_only=False)
    settings = get_settings(workbook['Settings']) if 'Settings' in workbook.sheetnames else {}
    workbook_manifest = {
        'workbook_path': str(WORKBOOK_PATH),
        'sheet_count': len(workbook.sheetnames),
        'generated_at': datetime.now(timezone.utc).isoformat(),
        'settings_snapshot': settings,
        'sheets': [],
    }
    workbook_review_rows = []

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        category = 'support' if sheet_name in SUPPORT_SHEETS else 'building'
        slug = slugify(sheet_name)
        raw_rows = extract_non_empty_rows(ws)

        review_flags: list[dict[str, Any]] = []
        normalized_payload: dict[str, Any]
        if category == 'building':
            normalized_payload, review_flags = extract_building_sheet(ws, settings)
        else:
            normalized_payload = {
                'rows': raw_rows,
            }

        structured_payload = {
            'workbook_path': str(WORKBOOK_PATH),
            'sheet_name': sheet_name,
            'sheet_slug': slug,
            'sheet_category': category,
            'scheme_name': canonical_scheme_name(sheet_name) if category == 'building' else None,
            'sheet_dimensions': {
                'max_row': ws.max_row,
                'max_column': ws.max_column,
            },
            'merged_ranges': [str(cell_range) for cell_range in ws.merged_cells.ranges],
            'settings_snapshot': settings,
            'raw_rows': raw_rows,
            'structured': normalized_payload,
            'review_flags': review_flags,
        }

        input_path = INPUT_DIR / f'{slug}.sheet.json'
        normalized_path = NORMALIZED_DIR / f'{slug}.normalized.json'
        write_json(input_path, structured_payload)
        write_json(normalized_path, {
            'sheet_name': sheet_name,
            'sheet_slug': slug,
            'sheet_category': category,
            'scheme_name': canonical_scheme_name(sheet_name) if category == 'building' else None,
            'settings_snapshot': settings,
            **normalized_payload,
            'review_flags': review_flags,
        })

        workbook_manifest['sheets'].append({
            'sheet_name': sheet_name,
            'sheet_slug': slug,
            'sheet_category': category,
            'input_json': str(input_path.relative_to(BASE_DIR)),
            'normalized_json': str(normalized_path.relative_to(BASE_DIR)),
            'review_flag_count': len(review_flags),
        })

        for flag in review_flags:
            workbook_review_rows.append({
                'sheet_name': sheet_name,
                'severity': flag.get('severity'),
                'type': flag.get('type'),
                'message': flag.get('message'),
                'row_index': flag.get('row_index'),
                'column_index': flag.get('column_index'),
                'legacy_label': flag.get('legacy_label'),
                'display_label': flag.get('display_label'),
                'reading_date_raw': flag.get('reading_date_raw'),
                'count': flag.get('count'),
            })

        print(f'Exported sheet JSON: {sheet_name}')

    write_json(REVIEW_DIR / 'workbook-sheet-export-manifest.json', workbook_manifest)
    write_json(REVIEW_DIR / 'workbook-review-summary.json', {
        'workbook_path': str(WORKBOOK_PATH),
        'generated_at': workbook_manifest['generated_at'],
        'total_flags': len(workbook_review_rows),
        'flags_by_sheet': Counter(row['sheet_name'] for row in workbook_review_rows),
        'flags_by_type': Counter(row['type'] for row in workbook_review_rows),
        'flags': workbook_review_rows,
    })
    write_review_csv(REVIEW_DIR / 'workbook-review-summary.csv', workbook_review_rows)


if __name__ == '__main__':
    main()