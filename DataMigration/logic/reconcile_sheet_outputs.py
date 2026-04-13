from __future__ import annotations

import csv
import io
import json
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from generate_workbook_sheet_summaries import (
    WORKBOOK_PASSWORD,
    WORKBOOK_PATH,
    build_categories,
    decrypt_workbook_bytes,
    infer_building_entities,
    infer_domains,
    infer_support_entities,
    load_json,
    normalize_text,
    summarize_attributes,
)


ROOT_DIR = Path(__file__).resolve().parents[2]
MANIFEST_PATH = ROOT_DIR / 'DataMigration' / 'outputs' / 'reviews' / 'workbook-sheet-export-manifest.json'
REVIEW_SUMMARY_JSON_PATH = ROOT_DIR / 'DataMigration' / 'outputs' / 'reviews' / 'workbook-review-summary.json'
REVIEW_SUMMARY_CSV_PATH = ROOT_DIR / 'DataMigration' / 'outputs' / 'reviews' / 'workbook-review-summary.csv'
RECONCILIATION_REPORT_PATH = ROOT_DIR / 'DataMigration' / 'outputs' / 'reviews' / 'reconciliation-report.json'
NORMALIZED_DIR = ROOT_DIR / 'DataMigration' / 'outputs' / 'sheet-normalized'
SUMMARY_DIR = ROOT_DIR / 'DataMigration' / 'Summaries'
BUILDINGS_DIR = ROOT_DIR / 'Buildings' / 'buildings'

SCHEME_ALIASES = {
    'the azores': 'Azores',
    'genesis': 'Genisis on Fairmount',
    'bonifay court': 'Bonifay',
    'vilino glen': 'VILLINO GLEN',
    'l montagne': 'La Montagne',
    'lmontagne': 'La Montagne',
    "l'montagne": 'La Montagne',
    'rivonia gate': 'Rivonia Gates',
    'queensgate': 'QueensGate',
}

UNIT_LABEL_RE = re.compile(r'^(?P<prefix>[A-Za-z]+)\s*(?P<number>\d+)(?P<suffix>[A-Za-z]?)$')
BUILDING_FILE_SIGNAL_RE = re.compile(r'meter|reading|electricity|elec|prepaid|bulk|utility|kwh', re.IGNORECASE)
SETTINGS_SECRET_MARKERS = {'agent login details', 'username', 'password'}


@dataclass
class MonthColumn:
    column_index: int
    display_label: str | None
    reading_date_raw: Any
    tariff_table: str | None


def canonical_scheme_name(sheet_name: str) -> str:
    normalized = normalize_text(sheet_name)
    return SCHEME_ALIASES.get(normalized, sheet_name.strip())


def load_reconciliation_workbooks(workbook_bytes: bytes):
    formula_stream = io.BytesIO(workbook_bytes)
    data_stream = io.BytesIO(workbook_bytes)
    formula_workbook = load_workbook(formula_stream, data_only=False, read_only=False)
    data_workbook = load_workbook(data_stream, data_only=True, read_only=False)
    return formula_workbook, data_workbook


def serialize_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat(sep=' ')
    if isinstance(value, date):
        return value.isoformat()
    return value


def parse_text_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text or text == '#########':
        return None
    normalized = text.replace(' ', '').replace(',', '')
    try:
        return float(normalized)
    except ValueError:
        return None


def try_parse_date_value(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if value is None:
        return None

    text = str(value).strip()
    if not text or text == '#########':
        return None

    try:
        return datetime.fromisoformat(text.replace('Z', '+00:00')).date().isoformat()
    except ValueError:
        pass

    formats = [
        '%Y-%m',
        '%Y-%m-%d',
        '%Y-%m-%d %H:%M:%S',
        '%Y/%m/%d',
        '%Y/%m/%d %H:%M:%S',
        '%d/%m/%Y',
        '%m/%d/%Y',
        '%d-%m-%Y',
        '%d-%b-%y',
        '%d %b %Y',
        '%d %B %Y',
        '%b-%y',
        '%B-%y',
    ]
    for fmt in formats:
        try:
            parsed = datetime.strptime(text, fmt)
            if fmt == '%Y-%m':
                return parsed.strftime('%Y-%m-01')
            if fmt in {'%b-%y', '%B-%y'}:
                return parsed.strftime('%Y-%m-01')
            return parsed.date().isoformat()
        except ValueError:
            continue

    return None


def parse_reading_date(raw_value: Any, display_label: str | None) -> tuple[str | None, str | None]:
    raw_date = try_parse_date_value(raw_value)
    if raw_date:
        return raw_date, None
    if raw_value in (None, ''):
        derived = try_parse_date_value(display_label)
        if derived:
            return derived, 'derived-from-display-label'
        return None, 'missing-date'

    if str(raw_value).strip() == '#########':
        derived = try_parse_date_value(display_label)
        if derived:
            return derived, 'derived-from-display-label'
        return None, 'placeholder-date'

    derived = try_parse_date_value(display_label)
    if derived:
        return derived, 'derived-from-display-label'
    return None, 'unparseable-date'


def get_settings(ws) -> dict[str, Any]:
    settings = {
        'vat_percent': None,
        'ignore_negative_common_property_amounts': None,
        'tolerance_months': None,
        'tolerance_percentage': None,
    }
    for row in ws.iter_rows(min_row=1, max_row=40, min_col=1, max_col=4, values_only=True):
        row_values = list(row)
        if len(row_values) < 3:
            continue
        description = row_values[2]
        value = row_values[1]
        if description == 'Value' and row_values[1] == 'Description':
            continue
        if description == 'Ignore Negative Common Proper Amounts':
            settings['ignore_negative_common_property_amounts'] = bool(value)
        elif description == 'Error Tolerance Percentage':
            settings['tolerance_percentage'] = parse_text_number(value)
        elif description == 'Number of Months to compare average':
            settings['tolerance_months'] = int(parse_text_number(value) or 0)
        elif row_values[1] == 'VAT':
            settings['vat_percent'] = parse_text_number(description)
    return settings


def get_month_columns(formula_ws, data_ws) -> list[MonthColumn]:
    columns: list[MonthColumn] = []
    column_index = 5
    while True:
        display_formula = formula_ws.cell(1, column_index).value
        display_value = data_ws.cell(1, column_index).value
        raw_date_formula = formula_ws.cell(2, column_index).value
        raw_date_value = data_ws.cell(2, column_index).value
        tariff_formula = formula_ws.cell(3, column_index).value
        tariff_value = data_ws.cell(3, column_index).value

        display_label = display_value if display_value not in (None, '') else display_formula
        raw_date = raw_date_value if raw_date_value not in (None, '') else raw_date_formula
        tariff_table = tariff_value if tariff_value not in (None, '') else tariff_formula

        if display_label in (None, '') and raw_date in (None, ''):
            break

        columns.append(
            MonthColumn(
                column_index=column_index,
                display_label=str(display_label).strip() if display_label is not None else None,
                reading_date_raw=raw_date,
                tariff_table=str(tariff_table).strip() if tariff_table is not None else None,
            )
        )
        column_index += 1
    return columns


def find_row_by_label(ws, label: str, max_rows: int) -> int | None:
    target = normalize_text(label)
    for row_index in range(1, max_rows + 1):
        cell_value = normalize_text(ws.cell(row_index, 1).value)
        if cell_value == target:
            return row_index
    return None


def classify_meter_type(label: str) -> str:
    normalized = str(label or '').replace('*', '').strip().upper()
    if not normalized:
        return 'UNIT'
    if normalized.startswith('BUL') or 'BULK' in normalized:
        return 'BULK'
    if normalized.startswith('COM'):
        return 'COMMON'
    common_tokens = [
        'GATE', 'GUARD', 'STORE', 'PUMP', 'POOL', 'LIFT', 'BOREHOLE', 'CLUBHOUSE',
        'OFFICE', 'LAUNDRY', 'LIGHT', 'LOBBY', 'SEWER', 'GARDEN', 'SPRINKLER', 'PUBLIC'
    ]
    if any(token in normalized for token in common_tokens):
        return 'COMMON'
    return 'UNIT'


def canonical_unit_label(prefix: str, number: str, suffix: str) -> str:
    padded = number.zfill(2) if len(number) < 2 else number
    return f'{prefix.upper()} {padded}{suffix.upper()}'.strip()


def derive_entity_reference(label: str, meter_type: str) -> dict[str, Any]:
    reference = {
        'entity_kind': meter_type.lower(),
        'legacy_label': label,
    }
    if meter_type == 'UNIT':
        match = UNIT_LABEL_RE.match(label.strip())
        if match:
            prefix = match.group('prefix')
            number = match.group('number')
            suffix = match.group('suffix') or ''
            reference.update({
                'label_prefix': prefix.upper(),
                'numeric_component': int(number),
                'suffix': suffix.upper() or None,
                'canonical_unit_label': canonical_unit_label(prefix, number, suffix),
            })
    return reference


def extract_non_empty_rows(formula_ws) -> list[dict[str, Any]]:
    rows = []
    for row_index in range(1, formula_ws.max_row + 1):
        cells = []
        for column_index in range(1, formula_ws.max_column + 1):
            value = formula_ws.cell(row_index, column_index).value
            if value in (None, ''):
                continue
            cells.append({
                'column_index': column_index,
                'column_letter': formula_ws.cell(row_index, column_index).column_letter,
                'value': serialize_value(value),
            })
        if cells:
            rows.append({'row_index': row_index, 'cells': cells})
    return rows


def sanitize_settings_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    sanitized = []
    for row in rows:
        text_values = [normalize_text(cell.get('value')) for cell in row.get('cells', []) if isinstance(cell.get('value'), str)]
        if any(value in SETTINGS_SECRET_MARKERS for value in text_values):
            continue
        sanitized.append(row)
    return sanitized


def locate_building_folder(scheme_name: str) -> Path | None:
    normalized_target = normalize_text(scheme_name)
    candidates = [BUILDINGS_DIR / scheme_name]
    alias_name = SCHEME_ALIASES.get(normalized_target)
    if alias_name:
        candidates.append(BUILDINGS_DIR / alias_name)
    for path in BUILDINGS_DIR.iterdir():
        if not path.is_dir():
            continue
        if normalize_text(path.name) == normalized_target:
            candidates.append(path)
        elif alias_name and normalize_text(path.name) == normalize_text(alias_name):
            candidates.append(path)
    seen = set()
    for candidate in candidates:
        candidate_key = str(candidate).lower()
        if candidate_key in seen:
            continue
        seen.add(candidate_key)
        if candidate.exists() and candidate.is_dir():
            return candidate
    return None


def build_source_evidence(scheme_name: str, unit_labels: list[str]) -> dict[str, Any]:
    folder = locate_building_folder(scheme_name)
    if folder is None:
        return {
            'folder_exists': False,
            'matched_folder': None,
            'total_file_count': 0,
            'relevant_file_count': 0,
            'file_type_counts': {},
            'top_level_folders': [],
            'matched_unit_mentions': [],
            'notes': ['No retained source folder was found under Buildings/buildings for this scheme.'],
        }

    files = [path for path in folder.rglob('*') if path.is_file()]
    file_type_counts = Counter(path.suffix.lower() or '<no-extension>' for path in files)
    relevant_files = [path for path in files if BUILDING_FILE_SIGNAL_RE.search(path.name)]
    normalized_unit_labels = {normalize_text(label): label for label in unit_labels}
    matched_mentions = []
    for path in files:
        normalized_name = normalize_text(path.stem)
        for unit_key, unit_label in normalized_unit_labels.items():
            if unit_key and unit_key in normalized_name:
                matched_mentions.append({
                    'unit_label': unit_label,
                    'file': str(path),
                })
                break
        if len(matched_mentions) >= 40:
            break

    return {
        'folder_exists': True,
        'matched_folder': str(folder),
        'total_file_count': len(files),
        'relevant_file_count': len(relevant_files),
        'file_type_counts': dict(sorted(file_type_counts.items())),
        'top_level_folders': sorted(item.name for item in folder.iterdir() if item.is_dir()),
        'matched_unit_mentions': matched_mentions,
        'notes': ['Cross-reference is based on retained file/folder names only; OCR/text extraction was not performed for PDFs or images.'],
    }


def build_review_rows(sheet_name: str, review_flags: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    for flag in review_flags:
        rows.append({
            'sheet_name': sheet_name,
            'severity': flag.get('severity'),
            'type': flag.get('type'),
            'message': flag.get('message'),
            'row_index': flag.get('row_index'),
            'column_index': flag.get('column_index'),
            'legacy_label': flag.get('legacy_label'),
            'display_label': flag.get('display_label'),
            'reading_date_raw': flag.get('reading_date_raw'),
            'count': flag.get('count'),
        })
    return rows


def write_review_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    fieldnames = [
        'sheet_name', 'severity', 'type', 'message', 'row_index', 'column_index',
        'legacy_label', 'display_label', 'reading_date_raw', 'count'
    ]
    with path.open('w', newline='', encoding='utf-8') as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def reconcile_building_sheet(sheet_record: dict[str, Any], formula_ws, data_ws, settings_snapshot: dict[str, Any]) -> tuple[dict[str, Any], dict[str, Any], list[dict[str, Any]]]:
    review_flags: list[dict[str, Any]] = []
    month_columns = get_month_columns(formula_ws, data_ws)
    max_rows = formula_ws.max_row
    header_scan_limit = min(max_rows, 250)
    electricity_header = find_row_by_label(formula_ws, 'ELECTRICITY', header_scan_limit)
    water_header = find_row_by_label(formula_ws, 'WATER', header_scan_limit)

    normalized_months = []
    derived_dates = 0
    for month in month_columns:
        reading_date, reading_date_note = parse_reading_date(month.reading_date_raw, month.display_label)
        if reading_date_note == 'derived-from-display-label':
            derived_dates += 1
        month_record = {
            'column_index': month.column_index,
            'display_label': month.display_label,
            'reading_date_raw': serialize_value(month.reading_date_raw),
            'reading_date': reading_date,
            'reading_date_note': reading_date_note,
            'tariff_table': month.tariff_table,
        }
        normalized_months.append(month_record)

        if not month.display_label:
            review_flags.append({
                'severity': 'medium',
                'type': 'missing-month-display-label',
                'column_index': month.column_index,
                'message': 'Month column is missing a display label.',
            })
        if reading_date is None:
            review_flags.append({
                'severity': 'high',
                'type': 'unresolved-reading-date',
                'column_index': month.column_index,
                'display_label': month.display_label,
                'reading_date_raw': serialize_value(month.reading_date_raw),
                'message': f'Reading date could not be resolved for month column {month.column_index}.',
            })

    charge_modes = []
    for row_index in range(4, 12):
        charge_name = formula_ws.cell(row_index, 1).value
        distribution_mode = data_ws.cell(row_index, 3).value
        if distribution_mode in (None, ''):
            distribution_mode = formula_ws.cell(row_index, 3).value
        if charge_name:
            charge_modes.append({
                'source_row': row_index,
                'charge_name': str(charge_name).strip(),
                'distribution_mode': str(distribution_mode).strip() if distribution_mode not in (None, '') else None,
            })
        else:
            review_flags.append({
                'severity': 'medium',
                'type': 'missing-charge-heading',
                'row_index': row_index,
                'message': 'Expected charge heading row is blank.',
            })

    if electricity_header is None:
        review_flags.append({
            'severity': 'high',
            'type': 'missing-electricity-header',
            'message': 'The sheet does not contain an ELECTRICITY section header.',
        })
        payload = {
            'sheet_name': sheet_record['sheet_name'],
            'sheet_slug': sheet_record['sheet_slug'],
            'sheet_category': 'building',
            'scheme_name': canonical_scheme_name(sheet_record['sheet_name']),
            'settings_snapshot': settings_snapshot,
            'month_columns': normalized_months,
            'charge_modes': charge_modes,
            'sections': {
                'electricity_header_row': None,
                'water_header_row': water_header,
            },
            'electricity_rows': [],
        }
        return payload, {'derived_dates': derived_dates, 'resolved_formula_values': 0, 'source_evidence': build_source_evidence(payload['scheme_name'], [])}, review_flags

    electricity_rows = []
    seen_labels = Counter()
    current_row = electricity_header + 1
    blank_streak = 0
    resolved_formula_values = 0
    unit_labels: list[str] = []

    while current_row <= max_rows:
        label = formula_ws.cell(current_row, 1).value
        if label is None:
            blank_streak += 1
            if blank_streak >= 12:
                break
            current_row += 1
            continue
        blank_streak = 0
        label_text = str(label).strip()
        upper_label = label_text.upper()

        if upper_label == 'WATER':
            break
        if upper_label in {'ELECTRICITY', 'UNIT'}:
            current_row += 1
            continue

        seen_labels[label_text] += 1
        meter_type = classify_meter_type(label_text)
        if meter_type == 'UNIT':
            unit_labels.append(label_text)
        prepaid_marker = data_ws.cell(current_row, 2).value
        if prepaid_marker in (None, ''):
            prepaid_marker = formula_ws.cell(current_row, 2).value
        pq_factor = parse_text_number(data_ws.cell(current_row, 3).value)
        if pq_factor is None:
            pq_factor = parse_text_number(formula_ws.cell(current_row, 3).value)

        readings = []
        placeholder_count = 0
        non_numeric_count = 0
        formula_without_cache_count = 0

        for month in normalized_months:
            formula_value = formula_ws.cell(current_row, month['column_index']).value
            cached_value = data_ws.cell(current_row, month['column_index']).value

            if formula_value in (None, '') and cached_value in (None, ''):
                continue

            raw_value = formula_value if formula_value not in (None, '') else cached_value
            reading_value = parse_text_number(cached_value)
            if reading_value is None:
                reading_value = parse_text_number(raw_value)

            formula_text = formula_value if isinstance(formula_value, str) and formula_value.startswith('=') else None
            value_source = 'literal'
            if formula_text:
                value_source = 'cached-formula-value' if parse_text_number(cached_value) is not None else 'formula-without-cached-value'
            elif reading_value is None and raw_value not in (None, ''):
                value_source = 'non-numeric-text'

            if raw_value == '#########':
                placeholder_count += 1
            if formula_text and parse_text_number(cached_value) is not None:
                resolved_formula_values += 1
            if formula_text and parse_text_number(cached_value) is None:
                formula_without_cache_count += 1
            if raw_value not in (None, '') and reading_value is None:
                non_numeric_count += 1

            readings.append({
                'column_index': month['column_index'],
                'reading_label': month['display_label'],
                'reading_date_raw': month['reading_date_raw'],
                'reading_date': month['reading_date'],
                'reading_date_note': month['reading_date_note'],
                'tariff_table': month['tariff_table'],
                'reading_value_raw': serialize_value(raw_value),
                'reading_value': reading_value,
                'reading_value_formula': formula_text,
                'reading_value_cached': serialize_value(cached_value) if formula_text else None,
                'reading_value_source': value_source,
            })

        if not any(reading.get('reading_value') is not None for reading in readings):
            review_flags.append({
                'severity': 'medium',
                'type': 'row-with-no-readings',
                'row_index': current_row,
                'legacy_label': label_text,
                'message': 'Electricity row has no captured reading values.',
            })
        if placeholder_count > 0:
            review_flags.append({
                'severity': 'high',
                'type': 'placeholder-reading-values',
                'row_index': current_row,
                'legacy_label': label_text,
                'count': placeholder_count,
                'message': 'Electricity row contains placeholder reading values.',
            })
        if non_numeric_count > 0:
            review_flags.append({
                'severity': 'medium',
                'type': 'non-numeric-reading-values',
                'row_index': current_row,
                'legacy_label': label_text,
                'count': non_numeric_count,
                'message': 'Electricity row contains non-numeric reading content that could not be resolved.',
            })
        if formula_without_cache_count > 0:
            review_flags.append({
                'severity': 'high',
                'type': 'formula-without-cached-value',
                'row_index': current_row,
                'legacy_label': label_text,
                'count': formula_without_cache_count,
                'message': 'Electricity row contains formulas without cached numeric values.',
            })

        electricity_rows.append({
            'source_row': current_row,
            'legacy_label': label_text,
            'meter_type': meter_type,
            'entity_reference': derive_entity_reference(label_text, meter_type),
            'prepaid_marker': serialize_value(prepaid_marker),
            'pq_factor': pq_factor,
            'readings': readings,
        })
        current_row += 1

    for legacy_label, count in seen_labels.items():
        if count > 1:
            review_flags.append({
                'severity': 'high',
                'type': 'duplicate-legacy-label',
                'legacy_label': legacy_label,
                'count': count,
                'message': 'Legacy label appears more than once in the ELECTRICITY section.',
            })

    payload = {
        'sheet_name': sheet_record['sheet_name'],
        'sheet_slug': sheet_record['sheet_slug'],
        'sheet_category': 'building',
        'scheme_name': canonical_scheme_name(sheet_record['sheet_name']),
        'settings_snapshot': settings_snapshot,
        'month_columns': normalized_months,
        'charge_modes': charge_modes,
        'sections': {
            'electricity_header_row': electricity_header,
            'water_header_row': water_header,
        },
        'electricity_rows': electricity_rows,
    }

    metrics = {
        'derived_dates': derived_dates,
        'resolved_formula_values': resolved_formula_values,
        'source_evidence': build_source_evidence(payload['scheme_name'], unit_labels),
    }
    return payload, metrics, review_flags


def reconcile_support_sheet(sheet_record: dict[str, Any], existing_normalized: dict[str, Any], settings_snapshot: dict[str, Any]) -> tuple[dict[str, Any], dict[str, Any], list[dict[str, Any]]]:
    raw_rows = existing_normalized.get('rows', [])
    if sheet_record['sheet_slug'] == 'settings':
        raw_rows = sanitize_settings_rows(raw_rows)
    payload = {
        'sheet_name': sheet_record['sheet_name'],
        'sheet_slug': sheet_record['sheet_slug'],
        'sheet_category': 'support',
        'scheme_name': None,
        'settings_snapshot': settings_snapshot,
        'rows': raw_rows,
    }
    metrics = {
        'derived_dates': 0,
        'resolved_formula_values': 0,
        'source_evidence': None,
    }
    return payload, metrics, []


def main() -> None:
    manifest = load_json(MANIFEST_PATH)
    previous_review_summary = load_json(REVIEW_SUMMARY_JSON_PATH) if REVIEW_SUMMARY_JSON_PATH.exists() else None
    previous_flags_by_sheet = previous_review_summary.get('flags_by_sheet', {}) if previous_review_summary else {}
    previous_total_flags = previous_review_summary.get('total_flags', 0) if previous_review_summary else 0

    workbook_bytes = decrypt_workbook_bytes(WORKBOOK_PATH, WORKBOOK_PASSWORD)
    formula_workbook, data_workbook = load_reconciliation_workbooks(workbook_bytes)
    settings_snapshot = get_settings(data_workbook['Settings']) if 'Settings' in data_workbook.sheetnames else {}

    updated_review_rows: list[dict[str, Any]] = []
    reconciliation_rows = []

    for sheet_record in manifest['sheets']:
        sheet_name = sheet_record['sheet_name']
        sheet_slug = sheet_record['sheet_slug']
        existing_normalized = load_json(NORMALIZED_DIR / f'{sheet_slug}.normalized.json')
        formula_ws = formula_workbook[sheet_name]
        data_ws = data_workbook[sheet_name]

        if sheet_record['sheet_category'] == 'building':
            payload, metrics, review_flags = reconcile_building_sheet(sheet_record, formula_ws, data_ws, settings_snapshot)
        else:
            payload, metrics, review_flags = reconcile_support_sheet(sheet_record, existing_normalized, settings_snapshot)

        raw_rows = payload.get('rows') or extract_non_empty_rows(formula_ws)
        payload_for_summary = {
            **payload,
            'review_flags': review_flags,
        }

        domains = infer_domains(sheet_name, raw_rows, payload_for_summary)
        if sheet_record['sheet_category'] == 'building':
            entities = infer_building_entities(payload_for_summary)
        else:
            entities = infer_support_entities(raw_rows)
        categories = build_categories(sheet_name, sheet_record['sheet_category'], payload_for_summary, domains)
        attributes = summarize_attributes(sheet_record['sheet_category'], payload_for_summary, raw_rows)
        summary_path = SUMMARY_DIR / f'{sheet_slug}.summary.json'
        existing_formula_summary = {'formula_cell_count': 0, 'formula_patterns': []}
        if summary_path.exists():
            existing_summary = load_json(summary_path)
            existing_formula_summary = {
                'formula_cell_count': existing_summary.get('formulas_and_connections', {}).get('formula_cell_count', 0),
                'formula_patterns': existing_summary.get('formulas_and_connections', {}).get('formula_patterns', [])[:50],
            }

        normalized_output = {
            **payload,
            'entities_and_meanings': entities,
            'categories_and_domains': categories,
            'attributes_and_time_series': attributes,
            'formulas_and_connections_summary': existing_formula_summary,
            'source_evidence': metrics['source_evidence'],
            'cross_check': {
                'reconciled_at': datetime.now(UTC).isoformat(),
                'source_workbook': str(WORKBOOK_PATH),
                'summary_file': str(summary_path),
                'previous_review_flag_count': previous_flags_by_sheet.get(sheet_name, 0),
                'current_review_flag_count': len(review_flags),
                'derived_date_count': metrics['derived_dates'],
                'resolved_formula_value_count': metrics['resolved_formula_values'],
            },
            'review_flags': review_flags,
        }

        (NORMALIZED_DIR / f'{sheet_slug}.normalized.json').write_text(
            json.dumps(normalized_output, indent=2, ensure_ascii=True),
            encoding='utf-8',
        )

        updated_review_rows.extend(build_review_rows(sheet_name, review_flags))
        reconciliation_rows.append({
            'sheet_name': sheet_name,
            'sheet_slug': sheet_slug,
            'sheet_category': sheet_record['sheet_category'],
            'previous_review_flag_count': previous_flags_by_sheet.get(sheet_name, 0),
            'current_review_flag_count': len(review_flags),
            'derived_date_count': metrics['derived_dates'],
            'resolved_formula_value_count': metrics['resolved_formula_values'],
            'formula_cell_count': existing_formula_summary['formula_cell_count'],
            'source_evidence_folder': metrics['source_evidence']['matched_folder'] if metrics['source_evidence'] else None,
            'source_evidence_exists': metrics['source_evidence']['folder_exists'] if metrics['source_evidence'] else None,
        })
        sheet_record['review_flag_count'] = len(review_flags)

        print(f'Reconciled sheet: {sheet_name}')

    review_summary = {
        'workbook_path': str(WORKBOOK_PATH),
        'generated_at': datetime.now(UTC).isoformat(),
        'total_flags': len(updated_review_rows),
        'flags_by_sheet': Counter(row['sheet_name'] for row in updated_review_rows),
        'flags_by_type': Counter(row['type'] for row in updated_review_rows),
        'flags': updated_review_rows,
    }
    REVIEW_SUMMARY_JSON_PATH.write_text(json.dumps(review_summary, indent=2, ensure_ascii=True), encoding='utf-8')
    write_review_csv(REVIEW_SUMMARY_CSV_PATH, updated_review_rows)

    manifest['generated_at'] = datetime.now(UTC).isoformat()
    manifest['settings_snapshot'] = settings_snapshot
    MANIFEST_PATH.write_text(json.dumps(manifest, indent=2, ensure_ascii=True), encoding='utf-8')

    reconciliation_report = {
        'generated_at': datetime.now(UTC).isoformat(),
        'source_workbook': str(WORKBOOK_PATH),
        'previous_total_flags': previous_total_flags,
        'current_total_flags': len(updated_review_rows),
        'flag_delta': len(updated_review_rows) - previous_total_flags,
        'sheets': reconciliation_rows,
        'sheets_without_building_folder': [
            row['sheet_name']
            for row in reconciliation_rows
            if row['sheet_category'] == 'building' and not row['source_evidence_exists']
        ],
    }
    RECONCILIATION_REPORT_PATH.write_text(json.dumps(reconciliation_report, indent=2, ensure_ascii=True), encoding='utf-8')


if __name__ == '__main__':
    main()