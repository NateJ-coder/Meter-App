from __future__ import annotations

import io
import json
import re
from collections import Counter
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import msoffcrypto
from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parents[2]
WORKBOOK_PATH = ROOT_DIR / 'legacy-ud' / 'workbook-exports' / 'Utility Dash 9 Mar 2026.xlsm'
WORKBOOK_PASSWORD = 'ud'
MANIFEST_PATH = ROOT_DIR / 'DataMigration' / 'outputs' / 'reviews' / 'workbook-sheet-export-manifest.json'
INPUT_DIR = ROOT_DIR / 'DataMigration' / 'inputs' / 'sheet-json'
NORMALIZED_DIR = ROOT_DIR / 'DataMigration' / 'outputs' / 'sheet-normalized'
SUMMARY_DIR = ROOT_DIR / 'DataMigration' / 'Summaries'
VBA_MODULES_DIR = ROOT_DIR / 'legacy-ud' / 'vba-source' / 'modules'
VBA_OBJECTS_DIR = ROOT_DIR / 'legacy-ud' / 'vba-source' / 'objects'

IMPORTANT_BUILDING_VBA = [
    'D_BreakDown_Electricity.bas',
    'D_Breakdown_Data_Load.bas',
    'I_CSV_Import.bas',
    'I_Importing.bas',
    'T_Tolorance.bas',
    'Z_SAVE_LOAD_DATA.bas',
    'ThisWorkbook.cls',
]

SUPPORT_VBA_HINTS = {
    'home': ['Home_Sheet.cls', 'C_Navigation.bas', 'ThisWorkbook.cls'],
    'settings': ['A_Declare_Load.bas', 'Z_SAVE_LOAD_DATA.bas', 'ThisWorkbook.cls'],
    'tariffs': ['B_Tariff_Original.bas', 'B_Tariff_Other.bas', 'B_Tarrif_New.bas', 'New_Tariff_Backup.bas'],
    'readablereport': ['B_Building_Report.bas'],
    'buildingreport': ['B_Building_Report.bas'],
    'readablereporttemplate': ['B_Building_Report.bas'],
    'invoice': ['E_Invoice.bas'],
    'wcu-output': ['G_WCU_Out.bas'],
    'bcm-output': ['F_BCM_Out.bas', 'F_BCM_Transfer.bas'],
    'elecbreakdown': ['D_BreakDown_Electricity.bas', 'D_Breakdown_Data_Load.bas', 'T_Tolorance.bas'],
    'waterbreakdown': ['D_Breakdown_Water.bas', 'D_Breakdown_Data_Load.bas'],
    'unithistory': ['B_Building_Report.bas', 'Z_SAVE_LOAD_DATA.bas'],
    'checklist': ['Other.bas'],
    'sheet2': ['Other.bas'],
}

DOMAIN_KEYWORDS = {
    'electricity': ['electricity', 'elec', 'bulk', 'meter'],
    'water': ['water', 'sanitation', 'refuse', 'wcu'],
    'charges': ['charge', 'levy', 'surcharge', 'common prop', 'network'],
    'reporting': ['report', 'invoice', 'history', 'output', 'summary'],
    'settings': ['settings', 'tolerance', 'vat', 'description', 'value'],
    'navigation': ['home', 'help', 'menu', 'navigation'],
}

FORMULA_REF_RE = re.compile(r'\$?[A-Z]{1,3}\$?\d+(?::\$?[A-Z]{1,3}\$?\d+)?')
PROCEDURE_RE = re.compile(r'^\s*(?:Public|Private|Friend)?\s*(?:Static\s+)?(?:Sub|Function)\s+([A-Za-z_][A-Za-z0-9_]*)', re.IGNORECASE)
UNIT_LABEL_RE = re.compile(r'^([A-Za-z]+)\s*([0-9]+[A-Za-z]?)')


def slugify(value: str) -> str:
    slug = re.sub(r'[^a-z0-9]+', '-', value.strip().lower())
    return slug.strip('-') or 'unknown'


def normalize_text(value: Any) -> str:
    return re.sub(r'[^a-z0-9]+', ' ', str(value or '').strip().lower()).strip()


def decrypt_workbook_bytes(path: Path, password: str) -> bytes:
    decrypted = io.BytesIO()
    with path.open('rb') as handle:
        office_file = msoffcrypto.OfficeFile(handle)
        office_file.load_key(password=password)
        office_file.decrypt(decrypted)
    return decrypted.getvalue()


def load_workbooks(workbook_bytes: bytes):
    formula_stream = io.BytesIO(workbook_bytes)
    data_stream = io.BytesIO(workbook_bytes)
    formula_workbook = load_workbook(formula_stream, data_only=False, read_only=True)
    data_workbook = load_workbook(data_stream, data_only=True, read_only=True)
    return formula_workbook, data_workbook


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding='utf-8'))


def build_cell_map(raw_rows: list[dict[str, Any]]) -> dict[tuple[int, int], Any]:
    cell_map: dict[tuple[int, int], Any] = {}
    for row in raw_rows:
        row_index = row['row_index']
        for cell in row.get('cells', []):
            cell_map[(row_index, cell['column_index'])] = cell.get('value')
    return cell_map


def extract_formula_references(formula: str) -> list[str]:
    return FORMULA_REF_RE.findall(formula)


def normalize_formula_pattern(formula: str) -> str:
    pattern = FORMULA_REF_RE.sub('<REF>', formula)
    pattern = re.sub(r'\b\d+(?:\.\d+)?\b', '<NUM>', pattern)
    return pattern


def extract_vba_index() -> dict[str, dict[str, Any]]:
    index: dict[str, dict[str, Any]] = {}
    for base_dir in (VBA_MODULES_DIR, VBA_OBJECTS_DIR):
        for path in sorted(base_dir.glob('*')):
            if not path.is_file():
                continue
            content = path.read_text(encoding='utf-8', errors='ignore')
            procedures = []
            for line in content.splitlines():
                match = PROCEDURE_RE.match(line)
                if match:
                    procedures.append(match.group(1))
            index[path.name] = {
                'path': str(path),
                'procedures': procedures,
            }
    return index


def related_vba_sources(sheet_name: str, sheet_slug: str, sheet_category: str, vba_index: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    names: list[str] = []
    if sheet_category == 'building':
        names.extend(IMPORTANT_BUILDING_VBA)
        if sheet_slug == 'the-azores':
            names.append('Azores_Sheet.cls')
    else:
        for key, file_names in SUPPORT_VBA_HINTS.items():
            if key in sheet_slug:
                names.extend(file_names)
        if not names:
            names.append('ThisWorkbook.cls')

    unique_names = []
    seen = set()
    for name in names:
        if name in vba_index and name not in seen:
            unique_names.append(name)
            seen.add(name)
    return [vba_index[name] for name in unique_names]


def infer_domains(sheet_name: str, raw_rows: list[dict[str, Any]], normalized_payload: dict[str, Any]) -> list[str]:
    text_values = [sheet_name, normalized_payload.get('sheet_category', '')]
    for row in raw_rows[:80]:
        for cell in row.get('cells', []):
            value = cell.get('value')
            if isinstance(value, str):
                text_values.append(value)
    haystack = '\n'.join(text_values).lower()
    domains = []
    for domain, keywords in DOMAIN_KEYWORDS.items():
        if any(keyword in haystack for keyword in keywords):
            domains.append(domain)
    return domains


def infer_building_entities(normalized_payload: dict[str, Any]) -> list[dict[str, Any]]:
    electricity_rows = normalized_payload.get('electricity_rows', [])
    unit_labels = [row['legacy_label'] for row in electricity_rows if row.get('meter_type') == 'UNIT']
    bulk_labels = [row['legacy_label'] for row in electricity_rows if row.get('meter_type') == 'BULK']
    common_labels = [row['legacy_label'] for row in electricity_rows if row.get('meter_type') == 'COMMON']

    prefix_counter = Counter()
    for label in unit_labels:
        match = UNIT_LABEL_RE.match(label)
        if match:
            prefix_counter[match.group(1).upper()] += 1

    entities = [
        {
            'entity_type': 'building',
            'examples': [normalized_payload.get('sheet_name'), normalized_payload.get('scheme_name')],
            'meaning': 'Primary scheme/building represented by the worksheet.',
        }
    ]

    if prefix_counter:
        dominant_prefix, dominant_count = prefix_counter.most_common(1)[0]
        entities.append({
            'entity_type': 'unit_identifier_pattern',
            'pattern': f'{dominant_prefix} <unit number><optional suffix>',
            'examples': unit_labels[:12],
            'meaning': f'Unit-level meter labels. Prefix {dominant_prefix} appears to abbreviate the building or scheme name and is followed by the unit number.',
            'occurrence_count': dominant_count,
        })

    if bulk_labels:
        entities.append({
            'entity_type': 'bulk_meter_pattern',
            'pattern': 'Bulk <number>',
            'examples': bulk_labels,
            'meaning': 'Bulk supply meter identifiers used for building-level supply or aggregation.',
        })

    if common_labels:
        entities.append({
            'entity_type': 'common_property_pattern',
            'pattern': 'Common-property/service labels',
            'examples': common_labels[:12],
            'meaning': 'Meters associated with shared infrastructure, common property, or service areas rather than private units.',
        })

    return entities


def infer_support_entities(raw_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    labels = []
    for row in raw_rows[:120]:
        row_values = [cell.get('value') for cell in row.get('cells', []) if isinstance(cell.get('value'), str)]
        if not row_values:
            continue
        labels.append(' | '.join(row_values[:3]))
    deduped = []
    seen = set()
    for label in labels:
        normalized = normalize_text(label)
        if not normalized or normalized in seen:
            continue
        deduped.append(label)
        seen.add(normalized)
        if len(deduped) >= 20:
            break
    return [{
        'entity_type': 'sheet_labels_and_headings',
        'examples': deduped,
        'meaning': 'Prominent labels/headings captured from the support sheet that define its business purpose and fields.',
    }]


def extract_formula_cells(sheet_name: str, formula_ws, data_ws, cell_map: dict[tuple[int, int], Any]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    formula_cells: list[dict[str, Any]] = []
    pattern_counter = Counter()
    for row in formula_ws.iter_rows():
        for cell in row:
            value = cell.value
            if not isinstance(value, str) or not value.startswith('='):
                continue
            cached_value = data_ws[cell.coordinate].value
            row_label = cell_map.get((cell.row, 1))
            heading_primary = cell_map.get((1, cell.column))
            heading_secondary = cell_map.get((2, cell.column))
            heading_tertiary = cell_map.get((3, cell.column))
            references = extract_formula_references(value)
            pattern = normalize_formula_pattern(value)
            pattern_counter[pattern] += 1
            formula_cells.append({
                'sheet_name': sheet_name,
                'cell': cell.coordinate,
                'row_index': cell.row,
                'column_index': cell.column,
                'row_label': row_label,
                'heading_primary': heading_primary,
                'heading_secondary': heading_secondary,
                'heading_tertiary': heading_tertiary,
                'formula': value,
                'cached_value': cached_value,
                'references': references,
            })
    formula_patterns = [
        {'pattern': pattern, 'count': count}
        for pattern, count in pattern_counter.most_common()
    ]
    return formula_cells, formula_patterns


def summarize_attributes(sheet_category: str, normalized_payload: dict[str, Any], raw_rows: list[dict[str, Any]]) -> dict[str, Any]:
    if sheet_category == 'building':
        month_columns = normalized_payload.get('month_columns', [])
        resolved = [item for item in month_columns if item.get('reading_date')]
        unresolved = [item for item in month_columns if not item.get('reading_date')]
        tariff_tables = sorted({item.get('tariff_table') for item in month_columns if item.get('tariff_table')})
        return {
            'time_series_type': 'month_columns',
            'month_columns': month_columns,
            'resolved_month_count': len(resolved),
            'unresolved_month_count': len(unresolved),
            'first_resolved_month': resolved[0] if resolved else None,
            'last_resolved_month': resolved[-1] if resolved else None,
            'tariff_tables': tariff_tables,
        }

    date_like_values = []
    for row in raw_rows:
        for cell in row.get('cells', []):
            value = cell.get('value')
            if not isinstance(value, str):
                continue
            if re.search(r'\b\d{4}[-/]\d{2}(?:[-/]\d{2})?\b', value) or re.search(r'\b\d{1,2}[./-]\d{1,2}[./-]\d{2,4}\b', value):
                date_like_values.append({
                    'row_index': row['row_index'],
                    'column_index': cell['column_index'],
                    'value': value,
                })
    return {
        'time_series_type': 'date_like_values',
        'date_like_values': date_like_values[:200],
        'top_rows': raw_rows[:20],
    }


def build_categories(sheet_name: str, sheet_category: str, normalized_payload: dict[str, Any], domains: list[str]) -> dict[str, Any]:
    categories = {
        'sheet_category': sheet_category,
        'domains': domains,
    }
    if sheet_category == 'building':
        categories.update({
            'meter_types': sorted({row.get('meter_type') for row in normalized_payload.get('electricity_rows', []) if row.get('meter_type')}),
            'charge_names': [item.get('charge_name') for item in normalized_payload.get('charge_modes', []) if item.get('charge_name')],
            'distribution_modes': [item.get('distribution_mode') for item in normalized_payload.get('charge_modes', []) if item.get('distribution_mode')],
            'sections': normalized_payload.get('sections', {}),
        })
    else:
        categories['sheet_name'] = sheet_name
    return categories


def summarize_sheet(
    sheet_record: dict[str, Any],
    input_payload: dict[str, Any],
    normalized_payload: dict[str, Any],
    formula_ws,
    data_ws,
    vba_index: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    raw_rows = input_payload.get('raw_rows', [])
    cell_map = build_cell_map(raw_rows)
    formula_cells, formula_patterns = extract_formula_cells(sheet_record['sheet_name'], formula_ws, data_ws, cell_map)
    domains = infer_domains(sheet_record['sheet_name'], raw_rows, normalized_payload)

    if sheet_record['sheet_category'] == 'building':
        entities = infer_building_entities(normalized_payload)
    else:
        entities = infer_support_entities(raw_rows)

    summary = {
        'summary_type': 'utility-dash-sheet-summary',
        'generated_at': datetime.now(UTC).isoformat(),
        'workbook': {
            'source_file': str(WORKBOOK_PATH),
            'source_file_is_encrypted': True,
            'decryption_password_used': True,
        },
        'sheet': {
            'sheet_name': sheet_record['sheet_name'],
            'sheet_slug': sheet_record['sheet_slug'],
            'sheet_category': sheet_record['sheet_category'],
            'summary_file_name': f"{sheet_record['sheet_slug']}.summary.json",
        },
        'entities_and_meanings': entities,
        'categories_and_domains': build_categories(sheet_record['sheet_name'], sheet_record['sheet_category'], normalized_payload, domains),
        'attributes_and_time_series': summarize_attributes(sheet_record['sheet_category'], normalized_payload, raw_rows),
        'payload': {
            'sheet_dimensions': input_payload.get('sheet_dimensions'),
            'merged_ranges': input_payload.get('merged_ranges'),
            'raw_rows': raw_rows,
            'normalized_payload': normalized_payload,
        },
        'formulas_and_connections': {
            'formula_cell_count': len(formula_cells),
            'formula_patterns': formula_patterns,
            'formula_cells': formula_cells,
        },
        'related_vba_sources': related_vba_sources(sheet_record['sheet_name'], sheet_record['sheet_slug'], sheet_record['sheet_category'], vba_index),
    }
    return summary


def main() -> None:
    SUMMARY_DIR.mkdir(parents=True, exist_ok=True)

    manifest = load_json(MANIFEST_PATH)
    vba_index = extract_vba_index()
    workbook_bytes = decrypt_workbook_bytes(WORKBOOK_PATH, WORKBOOK_PASSWORD)
    formula_workbook, data_workbook = load_workbooks(workbook_bytes)

    summary_manifest = {
        'generated_at': datetime.now(UTC).isoformat(),
        'workbook_source': str(WORKBOOK_PATH),
        'sheet_count': len(manifest['sheets']),
        'summaries': [],
    }

    for sheet_record in manifest['sheets']:
        sheet_name = sheet_record['sheet_name']
        sheet_slug = sheet_record['sheet_slug']
        input_payload = load_json(INPUT_DIR / f'{sheet_slug}.sheet.json')
        normalized_payload = load_json(NORMALIZED_DIR / f'{sheet_slug}.normalized.json')
        formula_ws = formula_workbook[sheet_name]
        data_ws = data_workbook[sheet_name]

        summary = summarize_sheet(sheet_record, input_payload, normalized_payload, formula_ws, data_ws, vba_index)
        output_path = SUMMARY_DIR / f'{sheet_slug}.summary.json'
        output_path.write_text(json.dumps(summary, indent=2, ensure_ascii=True), encoding='utf-8')

        summary_manifest['summaries'].append({
            'sheet_name': sheet_name,
            'sheet_slug': sheet_slug,
            'sheet_category': sheet_record['sheet_category'],
            'summary_path': str(output_path),
            'formula_cell_count': summary['formulas_and_connections']['formula_cell_count'],
            'domain_count': len(summary['categories_and_domains']['domains']),
        })

        print(f'Summarized sheet: {sheet_name}')

    (SUMMARY_DIR / 'summaries-manifest.json').write_text(
        json.dumps(summary_manifest, indent=2, ensure_ascii=True),
        encoding='utf-8',
    )


if __name__ == '__main__':
    main()