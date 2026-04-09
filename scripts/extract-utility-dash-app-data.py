from __future__ import annotations

import csv
import json
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


WORKBOOK_PATH = Path(r"c:\Projects\Meter App\legacy-ud\workbook-exports\macro-free\Utility Dash 9 Mar decrypted working copy (2) 2026.xlsx")
OUTPUT_DIR = Path(r"c:\Projects\Meter App\source-documents\03-extracted-outputs\utility-dash-app")

SUPPORT_SHEETS = {
    'CheckList', 'ReadableReport', 'BuildingReport', 'BuildingReportTemplate', 'Variable_List',
    'ReadableReportTemplate (3)', 'ReadableReportTemplate (2)', 'Home', 'Help', 'Settings', 'Tariffs',
    'UnitHistory', 'ElecBreakDown', 'WaterBreakdown', 'ReadableReportTemplate', 'ReadableReportTemplate Water',
    'Invoice', 'WCU Output', 'BCM Output', 'Sheet2'
}

SCHEME_ALIASES = {
    'the azores': 'Azores',
    'genesis': 'Genisis on Fairmount',
    'bonifay court': 'Bonifay',
    'vilino glen': 'VILLINO GLEN',
    "l montagne": 'La Montagne',
    "lmontagne": 'La Montagne',
    "l'montagne": 'La Montagne',
    'rivonia gate': 'Rivonia Gates',
    'queensgate': 'QueensGate',
}


@dataclass
class MonthColumn:
    column_index: int
    display_label: str | None
    reading_date_raw: Any
    tariff_table: str | None


def slugify(value: str) -> str:
    slug = re.sub(r'[^a-z0-9]+', '-', value.strip().lower())
    return slug.strip('-') or 'unknown'


def normalize_token(value: Any) -> str:
    return re.sub(r'[^a-z0-9]+', ' ', str(value or '').strip().lower()).strip()


def canonical_scheme_name(sheet_name: str) -> str:
    normalized = normalize_token(sheet_name)
    return SCHEME_ALIASES.get(normalized, sheet_name.strip())


def deterministic_id(prefix: str, *parts: Any) -> str:
    tokens = [slugify(str(part)) for part in parts if str(part or '').strip()]
    return f"{prefix}-{'-'.join(tokens)}"


def current_timestamp() -> str:
    return datetime.now(timezone.utc).isoformat()


def parse_text_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text or text == '#########':
        return None
    normalized = text.replace(' ', '').replace(',', '')
    try:
        return float(normalized)
    except ValueError:
        return None


def parse_reading_date(raw_value: Any, display_label: str | None) -> tuple[str | None, str | None]:
    if isinstance(raw_value, datetime):
        return raw_value.date().isoformat(), None

    if raw_value is None:
        return None, 'missing-date'

    text = str(raw_value).strip()
    if not text or text == '#########':
        return None, 'placeholder-date'

    formats = ['%Y-%m', '%Y/%m/%d', '%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y']
    for fmt in formats:
        try:
            parsed = datetime.strptime(text, fmt)
            if fmt == '%Y-%m':
                return parsed.strftime('%Y-%m-01'), None
            return parsed.date().isoformat(), None
        except ValueError:
            continue

    if display_label:
        try:
            parsed = datetime.strptime(display_label, '%b-%y')
            return parsed.strftime('%Y-%m-01'), 'derived-from-label'
        except ValueError:
            pass

    return None, 'unparseable-date'


def get_settings(ws) -> dict[str, Any]:
    settings = {
        'vat_percent': None,
        'ignore_negative_common_property_amounts': None,
        'tolerance_months': None,
        'tolerance_percentage': None,
    }
    for row in ws.iter_rows(min_row=1, max_row=40, min_col=1, max_col=4, values_only=True):
        row_values = list(row)
        if len(row_values) < 3:
            continue
        description = row_values[2]
        value = row_values[1]
        if description == 'Value' and row_values[1] == 'Description':
            continue
        if description == 'Ignore Negative Common Proper Amounts':
            settings['ignore_negative_common_property_amounts'] = bool(value)
        elif description == 'Error Tolerance Percentage':
            settings['tolerance_percentage'] = parse_text_number(value)
        elif description == 'Number of Months to compare average':
            settings['tolerance_months'] = int(parse_text_number(value) or 0)
        elif row_values[1] == 'VAT':
            settings['vat_percent'] = parse_text_number(description)
    return settings


def get_month_columns(ws) -> list[MonthColumn]:
    columns: list[MonthColumn] = []
    column_index = 5
    while True:
        raw_date = ws.cell(2, column_index).value
        display_label = ws.cell(1, column_index).value
        tariff_table = ws.cell(3, column_index).value
        if raw_date in (None, '') and display_label in (None, ''):
            break
        columns.append(MonthColumn(
            column_index=column_index,
            display_label=str(display_label).strip() if display_label is not None else None,
            reading_date_raw=raw_date,
            tariff_table=str(tariff_table).strip() if tariff_table is not None else None,
        ))
        column_index += 1
    return columns


def find_row_by_label(ws, label: str, max_rows: int) -> int | None:
    target = normalize_token(label)
    for row_index in range(1, max_rows + 1):
        cell_value = normalize_token(ws.cell(row_index, 1).value)
        if cell_value == target:
            return row_index
    return None


def classify_meter_type(label: str) -> str:
    normalized = str(label or '').replace('*', '').strip().upper()
    if not normalized:
        return 'UNIT'
    if normalized.startswith('BUL') or 'BULK' in normalized:
        return 'BULK'
    if normalized.startswith('COM'):
        return 'COMMON'
    common_tokens = ['GATE', 'GUARD', 'STORE', 'PUMP', 'POOL', 'LIFT', 'BOREHOLE', 'CLUBHOUSE', 'OFFICE', 'LAUNDRY', 'LIGHT', 'LOBBY', 'SEWER', 'GARDEN', 'SPRINKLER']
    if any(token in normalized for token in common_tokens):
        return 'COMMON'
    return 'UNIT'


def get_consumption(current_value: float, previous_value: float) -> float:
    current_len = len(str(int(current_value))) if float(current_value).is_integer() else len(str(current_value))
    previous_len = len(str(int(previous_value))) if float(previous_value).is_integer() else len(str(previous_value))
    previous_text = str(int(previous_value)) if float(previous_value).is_integer() else str(previous_value)
    first_num = previous_text[:1]
    if current_len < previous_len and first_num == '9':
        return float(current_value) - float(previous_value) + (10 ** (previous_len - 1))
    return float(current_value) - float(previous_value)


def compute_tolerance_flag(consumptions: list[float], current_index: int, tolerance_months: int, tolerance_percentage: float) -> dict[str, Any] | None:
    if tolerance_months <= 0 or tolerance_percentage is None:
        return None
    if current_index < tolerance_months:
        return None
    current_consumption = consumptions[current_index]
    previous_window = consumptions[current_index - tolerance_months:current_index]
    if len(previous_window) < tolerance_months:
        return None
    past_average = int(sum(previous_window) / tolerance_months)
    tolerance_min = int(past_average * (1 - tolerance_percentage))
    tolerance_max = int(past_average * (1 + tolerance_percentage))
    if current_consumption < tolerance_max and current_consumption > tolerance_min:
        return None
    return {
        'type': 'tolerance',
        'severity': 'medium',
        'message': f'Consumption {current_consumption:.2f} fell outside tolerance band {tolerance_min} to {tolerance_max}.',
        'details': {
            'past_average': past_average,
            'tolerance_months': tolerance_months,
            'tolerance_percentage': tolerance_percentage,
            'tolerance_min': tolerance_min,
            'tolerance_max': tolerance_max,
        }
    }


def extract_sheet_rows(ws, month_columns: list[MonthColumn], settings: dict[str, Any]) -> dict[str, Any]:
    max_rows = ws.max_row
    header_scan_limit = min(max_rows, 250)
    electricity_header = find_row_by_label(ws, 'ELECTRICITY', header_scan_limit)
    water_header = find_row_by_label(ws, 'WATER', header_scan_limit)
    if electricity_header is None:
        return {'inventory_rows': [], 'reading_rows': [], 'unresolved_rows': [], 'charge_modes': []}

    charge_modes = []
    for row_index in range(4, 12):
        label = ws.cell(row_index, 1).value
        distribution_mode = ws.cell(row_index, 3).value
        if label:
            charge_modes.append({
                'charge_name': str(label).strip(),
                'distribution_mode': str(distribution_mode).strip() if distribution_mode is not None else None,
            })

    inventory_rows = []
    reading_rows = []
    unresolved_rows = []

    current_row = electricity_header + 1
    blank_streak = 0
    while current_row <= max_rows:
        label = ws.cell(current_row, 1).value
        if label is None:
            blank_streak += 1
            if blank_streak >= 12:
                break
            current_row += 1
            continue
        blank_streak = 0
        label_text = str(label).strip()
        upper_label = label_text.upper()
        if upper_label == 'WATER':
            break
        if upper_label in {'ELECTRICITY', 'UNIT'}:
            current_row += 1
            continue

        meter_type = classify_meter_type(label_text)
        prepaid_marker = ws.cell(current_row, 2).value
        pq_factor = parse_text_number(ws.cell(current_row, 3).value)
        inventory_rows.append({
            'legacy_label': label_text,
            'meter_type': meter_type,
            'prepaid_marker': str(prepaid_marker).strip() if prepaid_marker not in (None, '') else '',
            'pq_factor': pq_factor,
        })

        ordered_values: list[tuple[int, float, MonthColumn]] = []
        for month_index, month in enumerate(month_columns):
            reading_value = parse_text_number(ws.cell(current_row, month.column_index).value)
            if reading_value is None:
                continue
            ordered_values.append((month_index, reading_value, month))

        consumptions: list[float] = []
        for ordered_index, (month_index, reading_value, month) in enumerate(ordered_values):
            previous_reading_value = ordered_values[ordered_index - 1][1] if ordered_index > 0 else None
            consumption = None
            if previous_reading_value is not None:
                consumption = get_consumption(reading_value, previous_reading_value)
                prepaid_marker_text = str(prepaid_marker or '')
                if '*' in prepaid_marker_text:
                    factor_text = prepaid_marker_text.split('*', 1)[1]
                    factor_value = parse_text_number(factor_text)
                    if factor_value is not None:
                        consumption *= factor_value
                if consumption < 0:
                    consumption += 1000000
                consumptions.append(consumption)

            iso_date, date_note = parse_reading_date(month.reading_date_raw, month.display_label)
            tolerance_flag = None
            if consumption is not None:
                tolerance_flag = compute_tolerance_flag(
                    consumptions,
                    len(consumptions) - 1,
                    settings.get('tolerance_months') or 0,
                    settings.get('tolerance_percentage')
                )

            record = {
                'legacy_label': label_text,
                'meter_type': meter_type,
                'reading_label': month.display_label,
                'reading_date': iso_date,
                'reading_date_note': date_note,
                'tariff_table': month.tariff_table,
                'reading_value': reading_value,
                'previous_reading_value': previous_reading_value,
                'consumption': consumption,
                'flags': [tolerance_flag] if tolerance_flag else [],
            }

            if iso_date is None:
                unresolved_rows.append(record)
            else:
                reading_rows.append(record)

        current_row += 1

    return {
        'inventory_rows': inventory_rows,
        'reading_rows': reading_rows,
        'unresolved_rows': unresolved_rows,
        'charge_modes': charge_modes,
        'water_rows_found': max(0, (max_rows - (water_header or max_rows))) if water_header else 0,
    }


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open('w', newline='', encoding='utf-8') as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(WORKBOOK_PATH, data_only=False, read_only=False)
    settings = get_settings(workbook['Settings'])

    schemes = []
    buildings = []
    units = []
    meters = []
    cycles_by_key: dict[tuple[str, str], dict[str, Any]] = {}
    readings = []
    legacy_meter_map = []
    unresolved_dates = []
    summaries = []

    for sheet_name in workbook.sheetnames:
        if sheet_name in SUPPORT_SHEETS:
            continue

        print(f'Processing sheet: {sheet_name}')

        ws = workbook[sheet_name]
        scheme_name = canonical_scheme_name(sheet_name)
        scheme_id = deterministic_id('scheme', scheme_name)
        building_id = deterministic_id('building', scheme_name, 'main')

        if not any(item['id'] == scheme_id for item in schemes):
            schemes.append({
                'id': scheme_id,
                'name': scheme_name,
                'created_at': current_timestamp(),
                'imported_from': 'utility_dash_macrofree',
                'source_reference': sheet_name,
            })

        if not any(item['id'] == building_id for item in buildings):
            buildings.append({
                'id': building_id,
                'scheme_id': scheme_id,
                'name': scheme_name,
                'created_at': current_timestamp(),
                'imported_from': 'utility_dash_macrofree',
                'source_reference': sheet_name,
            })

        month_columns = get_month_columns(ws)
        extracted = extract_sheet_rows(ws, month_columns, settings)

        inventory_rows = extracted['inventory_rows']
        reading_rows = extracted['reading_rows']
        unresolved_rows = extracted['unresolved_rows']

        seen_units_for_sheet = set()
        seen_meters_for_sheet = set()

        for row in inventory_rows:
            legacy_label = row['legacy_label']
            meter_type = row['meter_type']
            unit_id = None
            unit_number = None

            if meter_type == 'UNIT':
                unit_number = legacy_label
                unit_id = deterministic_id('unit', scheme_name, unit_number)
                if unit_id not in seen_units_for_sheet:
                    seen_units_for_sheet.add(unit_id)
                    units.append({
                        'id': unit_id,
                        'building_id': building_id,
                        'unit_number': unit_number,
                        'status': 'occupied',
                        'created_at': current_timestamp(),
                        'imported_from': 'utility_dash_macrofree',
                        'source_reference': f'{sheet_name}:{legacy_label}',
                    })

            meter_id = deterministic_id('meter', scheme_name, meter_type, legacy_label)
            if meter_id in seen_meters_for_sheet:
                continue
            seen_meters_for_sheet.add(meter_id)

            meters.append({
                'id': meter_id,
                'scheme_id': scheme_id,
                'unit_id': unit_id,
                'meter_number': legacy_label,
                'meter_type': meter_type,
                'meter_role': 'unit' if meter_type == 'UNIT' else ('bulk' if meter_type == 'BULK' else 'common_property'),
                'service_type': 'electricity',
                'location_description': scheme_name if meter_type != 'UNIT' else unit_number,
                'last_reading': None,
                'last_reading_date': None,
                'pq_factor': row['pq_factor'],
                'prepaid_marker': row['prepaid_marker'],
                'created_at': current_timestamp(),
                'imported_from': 'utility_dash_macrofree',
                'source_reference': f'{sheet_name}:{legacy_label}',
                'source_confidence': 'medium',
            })

            legacy_meter_map.append({
                'id': deterministic_id('legacy-map', scheme_name, legacy_label),
                'legacy_label': legacy_label,
                'source_sheet': sheet_name,
                'meter_id': meter_id,
                'mapping_confidence': 'medium',
                'review_status': 'pending',
                'created_at': current_timestamp(),
            })

        meter_index = {item['meter_number']: item for item in meters if item['scheme_id'] == scheme_id}

        for row in reading_rows:
            meter = meter_index.get(row['legacy_label'])
            if not meter:
                continue
            cycle_key = (scheme_id, row['reading_date'])
            if cycle_key not in cycles_by_key:
                cycles_by_key[cycle_key] = {
                    'id': deterministic_id('cycle', scheme_name, row['reading_date']),
                    'scheme_id': scheme_id,
                    'name': f'{scheme_name} {row["reading_date"]}',
                    'start_date': row['reading_date'],
                    'end_date': row['reading_date'],
                    'status': 'CLOSED',
                    'created_at': current_timestamp(),
                    'imported_from': 'utility_dash_macrofree',
                    'source_reference': sheet_name,
                }

            reading_id = deterministic_id('reading', scheme_name, row['legacy_label'], row['reading_date'])
            readings.append({
                'id': reading_id,
                'meter_id': meter['id'],
                'cycle_id': cycles_by_key[cycle_key]['id'],
                'reading_date': row['reading_date'],
                'reading_value': row['reading_value'],
                'previous_reading': row['previous_reading_value'],
                'consumption': row['consumption'],
                'reading_type': 'actual',
                'capture_method': 'imported_excel',
                'review_status': 'pending' if row['flags'] else 'approved',
                'validation_status': 'needs_review' if row['flags'] else 'validated',
                'validation_reason': ' | '.join(flag['message'] for flag in row['flags']) if row['flags'] else '',
                'flags': row['flags'],
                'source_file': str(WORKBOOK_PATH),
                'source_row_reference': f'{sheet_name}:{row["legacy_label"]}:{row["reading_label"]}',
                'imported_from': 'utility_dash_macrofree',
                'created_at': current_timestamp(),
            })
            meter['last_reading'] = row['reading_value']
            meter['last_reading_date'] = row['reading_date']

        for row in unresolved_rows:
            unresolved_dates.append({
                'scheme_name': scheme_name,
                'source_sheet': sheet_name,
                'legacy_label': row['legacy_label'],
                'meter_type': row['meter_type'],
                'reading_label': row['reading_label'],
                'reading_date_note': row['reading_date_note'],
                'reading_value': row['reading_value'],
                'previous_reading_value': row['previous_reading_value'],
                'consumption': row['consumption'],
            })

        summaries.append({
            'scheme_name': scheme_name,
            'source_sheet': sheet_name,
            'months_available': len(month_columns),
            'inventory_meter_rows': len(inventory_rows),
            'resolved_readings': len(reading_rows),
            'unresolved_date_rows': len(unresolved_rows),
            'charge_modes': extracted['charge_modes'],
        })

    cycles = list(cycles_by_key.values())

    payload = {
        'metadata': {
            'source_file': str(WORKBOOK_PATH),
            'generated_at': current_timestamp(),
            'settings': settings,
            'notes': [
                'This payload is app-native and derived from the macro-free Utility Dash workbook.',
                'Readings with unresolved or placeholder dates were excluded from app-native reading import and written to unresolved-date review CSV.',
                'Consumption uses the workbook GetConsumption rollover rule.',
                'Tolerance flags use workbook Settings lookback and tolerance values.',
            ],
        },
        'schemes': schemes,
        'buildings': buildings,
        'units': units,
        'meters': meters,
        'cycles': cycles,
        'readings': readings,
        'legacy_meter_map': legacy_meter_map,
    }

    (OUTPUT_DIR / 'utility-dash-app-payload.json').write_text(json.dumps(payload, indent=2), encoding='utf-8')
    (OUTPUT_DIR / 'utility-dash-app-summary.json').write_text(json.dumps(summaries, indent=2), encoding='utf-8')

    write_csv(
        OUTPUT_DIR / 'utility-dash-app-meter-register.csv',
        [{
            'meter_id': meter['id'],
            'scheme_name': next(item['name'] for item in schemes if item['id'] == meter['scheme_id']),
            'service_type': meter['service_type'],
            'meter_number': meter['meter_number'],
            'meter_role': meter['meter_role'],
            'unit_number': next((unit['unit_number'] for unit in units if unit['id'] == meter['unit_id']), ''),
            'location_description': meter['location_description'],
            'parent_meter_id': meter.get('parent_meter_id') or '',
            'hierarchy_level': meter.get('hierarchy_level') or (0 if meter['meter_type'] == 'BULK' else 1),
            'reconciliation_group': meter.get('reconciliation_group') or meter['scheme_id'],
            'is_active': meter.get('is_active', True),
            'start_date': '',
            'end_date': '',
            'source_confidence': meter.get('source_confidence', 'medium'),
            'verified_by': '',
            'notes': '',
            'source_reference': meter['source_reference'],
        } for meter in meters],
        ['meter_id', 'scheme_name', 'service_type', 'meter_number', 'meter_role', 'unit_number', 'location_description', 'parent_meter_id', 'hierarchy_level', 'reconciliation_group', 'is_active', 'start_date', 'end_date', 'source_confidence', 'verified_by', 'notes', 'source_reference']
    )

    write_csv(
        OUTPUT_DIR / 'utility-dash-app-historical-readings.csv',
        [{
            'reading_id': reading['id'],
            'meter_id': reading['meter_id'],
            'reading_date': reading['reading_date'],
            'reading_value': reading['reading_value'],
            'reading_type': reading['reading_type'],
            'captured_by': '',
            'capture_method': reading['capture_method'],
            'source_file': reading['source_file'],
            'source_row_reference': reading['source_row_reference'],
            'evidence_link': '',
            'is_validated': True,
            'validation_status': reading['validation_status'],
            'validation_reason': reading['validation_reason'],
            'notes': '',
        } for reading in readings],
        ['reading_id', 'meter_id', 'reading_date', 'reading_value', 'reading_type', 'captured_by', 'capture_method', 'source_file', 'source_row_reference', 'evidence_link', 'is_validated', 'validation_status', 'validation_reason', 'notes']
    )

    write_csv(
        OUTPUT_DIR / 'utility-dash-app-unresolved-reading-dates.csv',
        unresolved_dates,
        ['scheme_name', 'source_sheet', 'legacy_label', 'meter_type', 'reading_label', 'reading_date_note', 'reading_value', 'previous_reading_value', 'consumption']
    )

    print(json.dumps({
        'workbook': str(WORKBOOK_PATH),
        'output_dir': str(OUTPUT_DIR),
        'schemes': len(schemes),
        'buildings': len(buildings),
        'units': len(units),
        'meters': len(meters),
        'cycles': len(cycles),
        'readings': len(readings),
        'legacy_meter_map': len(legacy_meter_map),
        'unresolved_date_rows': len(unresolved_dates),
    }, indent=2))


if __name__ == '__main__':
    main()
